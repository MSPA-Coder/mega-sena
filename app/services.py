from __future__ import annotations

import logging
import math
import secrets
import unicodedata
from collections import Counter
from datetime import date, datetime
from functools import lru_cache
from itertools import combinations
from pathlib import Path
from typing import BinaryIO, Iterable

import numpy as np
from openpyxl import load_workbook
from sqlalchemy import func

from . import db
from .models import Config, Draw, GeneratedBet

_log = logging.getLogger(__name__)


DEFAULT_CONFIG = {
    "bet_quantity": "6",
    "generation_amount": "5",
    "consecutive_count": "",
    "even_min": "",
    "even_max": "",
    "sum_min": "",
    "sum_max": "",
    "range_min_occupied": "",
    "range_max_per_band": "",
}

CONFIG_LIMITS = {
    "bet_quantity": (6, 15),
    "generation_amount": (1, 100),
    "consecutive_count": (0, 6),
    "even_min": (0, 6),
    "even_max": (0, 6),
    "sum_min": (0, None),
    "sum_max": (0, None),
    "range_min_occupied": (1, 6),
    "range_max_per_band": (1, 6),
}


def _norm(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return " ".join(text.lower().replace("_", " ").replace("-", " ").split())


def _parse_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return None
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _to_int(value: object) -> int | None:
    try:
        if value is None or value == "":
            return None
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _clamp_int(value: int, min_value: int | None = None, max_value: int | None = None) -> int:
    if min_value is not None:
        value = max(min_value, value)
    if max_value is not None:
        value = min(max_value, value)
    return value


def _clean_config_value(key: str, value: object) -> str:
    parsed = _to_int(value)
    default_value = DEFAULT_CONFIG[key]
    if parsed is None:
        return default_value if key in {"bet_quantity", "generation_amount"} else ""
    min_value, max_value = CONFIG_LIMITS[key]
    return str(_clamp_int(parsed, min_value, max_value))


def _normalize_config_values(values: dict[str, object]) -> dict[str, str]:
    normalized = {key: _clean_config_value(key, values.get(key, DEFAULT_CONFIG[key])) for key in DEFAULT_CONFIG}
    for key in ("consecutive_count", "even_min", "even_max"):
        if normalized[key]:
            normalized[key] = str(_clamp_int(int(normalized[key]), 0, 6))
    for key in ("range_min_occupied", "range_max_per_band"):
        if normalized[key]:
            normalized[key] = str(_clamp_int(int(normalized[key]), 1, 6))
    if normalized["even_min"] and normalized["even_max"] and int(normalized["even_min"]) > int(normalized["even_max"]):
        normalized["even_max"] = normalized["even_min"]
    if normalized["sum_min"] and normalized["sum_max"] and int(normalized["sum_min"]) > int(normalized["sum_max"]):
        normalized["sum_min"], normalized["sum_max"] = normalized["sum_max"], normalized["sum_min"]
    return normalized


def ensure_default_config() -> dict[str, str]:
    existing = {row.key: row for row in Config.query.all()}
    changed = False
    for key, value in DEFAULT_CONFIG.items():
        if key not in existing:
            db.session.add(Config(key=key, value=value))
            changed = True
    if changed:
        db.session.commit()
    return get_config_values()


def get_config_values() -> dict[str, str]:
    rows = {row.key: row.value for row in Config.query.all()}
    missing = [key for key in DEFAULT_CONFIG if key not in rows]
    if missing:
        for key in missing:
            db.session.add(Config(key=key, value=DEFAULT_CONFIG[key]))
        db.session.commit()
        rows = {row.key: row.value for row in Config.query.all()}
    return _normalize_config_values({key: rows.get(key, DEFAULT_CONFIG[key]) for key in DEFAULT_CONFIG})


def update_config_values(values: dict[str, object]) -> dict[str, str]:
    normalized = _normalize_config_values(values)
    rows = {row.key: row for row in Config.query.all()}
    for key, value in normalized.items():
        if key in rows:
            rows[key].value = value
        else:
            db.session.add(Config(key=key, value=value))
    db.session.commit()
    return normalized


def get_generation_defaults() -> dict[str, int | None]:
    values = get_config_values()
    return {
        "bet_quantity": int(values["bet_quantity"]),
        "generation_amount": int(values["generation_amount"]),
        "consecutive_count": _to_int(values["consecutive_count"]),
        "even_min": _to_int(values["even_min"]),
        "even_max": _to_int(values["even_max"]),
        "sum_min": _to_int(values["sum_min"]),
        "sum_max": _to_int(values["sum_max"]),
        "range_min_occupied": _to_int(values["range_min_occupied"]),
        "range_max_per_band": _to_int(values["range_max_per_band"]),
    }


def _money_to_cents(value: object) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(round(float(value) * 100))
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if not text:
        return 0
    text = text.replace(".", "").replace(",", ".")
    try:
        return int(round(float(text) * 100))
    except ValueError:
        return 0


def import_results_from_xlsx(source: str | Path | BinaryIO) -> dict[str, int]:
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        _log.error("Falha ao abrir planilha: %s", exc)
        raise RuntimeError(f"Não foi possível ler o arquivo: {exc}") from exc

    sheet = workbook[workbook.sheetnames[0]]
    if hasattr(sheet, "reset_dimensions"):
        sheet.reset_dimensions()
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return {"imported": 0, "updated": 0, "ignored": 0}

    normalized = [_norm(c) for c in header]

    def find_one(candidates: Iterable[str]) -> int | None:
        cand = {_norm(c) for c in candidates}
        for idx, name in enumerate(normalized):
            if name in cand:
                return idx
        for idx, name in enumerate(normalized):
            if any(c in name for c in cand):
                return idx
        return None

    contest_idx = find_one(["concurso", "contest", "numero concurso", "n concurso"])
    date_idx = find_one(["data sorteio", "data", "draw date"])
    winners_6_idx = find_one(["ganhadores 6 acertos", "ganhadores sena", "sena"])
    winners_5_idx = find_one(["ganhadores 5 acertos", "ganhadores quina", "quina"])
    winners_4_idx = find_one(["ganhadores 4 acertos", "ganhadores quadra", "quadra"])
    prize_idx = find_one(["rateio 6 acertos", "premio", "prêmio"])
    accumulated_idx = find_one(["acumulado 6 acertos", "acumulado"])
    quina_rateio_idx = find_one(["rateio 5 acertos", "rateio quina"])
    quadra_rateio_idx = find_one(["rateio 4 acertos", "rateio quadra"])

    number_indexes: list[int] = []
    for token in ["bola 1", "bola1", "dezena 1", "n1", "bola 2", "bola2", "dezena 2", "n2", "bola 3", "bola3", "dezena 3", "n3", "bola 4", "bola4", "dezena 4", "n4", "bola 5", "bola5", "dezena 5", "n5", "bola 6", "bola6", "dezena 6", "n6"]:
        idx = find_one([token])
        if idx is not None and idx not in number_indexes:
            number_indexes.append(idx)
        if len(number_indexes) == 6:
            break

    if contest_idx is None or len(number_indexes) < 6:
        return {"imported": 0, "updated": 0, "ignored": sum(1 for _ in rows)}

    imported = updated = ignored = 0
    existing_draws = {draw.contest: draw for draw in Draw.query.all()}
    seen_contests: set[int] = set()
    for row in rows:
        contest = _to_int(row[contest_idx] if contest_idx < len(row) else None)
        numbers = [_to_int(row[i] if i < len(row) else None) for i in number_indexes[:6]]
        if not contest or any(n is None or n < 1 or n > 60 for n in numbers) or len(set(numbers)) != 6:
            ignored += 1
            continue
        if contest in seen_contests:
            ignored += 1
            continue
        numbers = sorted(numbers)  # type: ignore[arg-type]
        derived = draw_parameters(numbers)
        payload = {
            "draw_date": _parse_date(row[date_idx]) if date_idx is not None and date_idx < len(row) else None,
            "n1": numbers[0], "n2": numbers[1], "n3": numbers[2],
            "n4": numbers[3], "n5": numbers[4], "n6": numbers[5],
            "total_sum": derived["total_sum"],
            "even_count": derived["even_count"],
            "consecutive_count": derived["consecutive_count"],
            "winners_6": _to_int(row[winners_6_idx]) if winners_6_idx is not None and winners_6_idx < len(row) else 0,
            "winners_5": _to_int(row[winners_5_idx]) if winners_5_idx is not None and winners_5_idx < len(row) else 0,
            "winners_4": _to_int(row[winners_4_idx]) if winners_4_idx is not None and winners_4_idx < len(row) else 0,
            "prize_cents": _money_to_cents(row[prize_idx]) if prize_idx is not None and prize_idx < len(row) else 0,
            "accumulated_cents": _money_to_cents(row[accumulated_idx]) if accumulated_idx is not None and accumulated_idx < len(row) else 0,
            "quina_rateio_cents": _money_to_cents(row[quina_rateio_idx]) if quina_rateio_idx is not None and quina_rateio_idx < len(row) else 0,
            "quadra_rateio_cents": _money_to_cents(row[quadra_rateio_idx]) if quadra_rateio_idx is not None and quadra_rateio_idx < len(row) else 0,
        }
        payload = {key: (0 if value is None and key.startswith("winners_") else value) for key, value in payload.items()}
        draw = existing_draws.get(contest)
        if draw:
            if all(getattr(draw, key) == value for key, value in payload.items()):
                ignored += 1
                seen_contests.add(contest)
                continue
            for key, value in payload.items():
                setattr(draw, key, value)
            updated += 1
        else:
            draw = Draw(contest=contest, **payload)
            db.session.add(draw)
            existing_draws[contest] = draw
            imported += 1
        seen_contests.add(contest)
    db.session.commit()
    _log.info("Importação concluída: %d novos, %d atualizados, %d ignorados.", imported, updated, ignored)
    return {"imported": imported, "updated": updated, "ignored": ignored}


def all_draw_numbers() -> list[list[int]]:
    return [d.numbers for d in Draw.query.order_by(Draw.contest).all()]


def count_even_numbers(numbers: Iterable[int]) -> int:
    return sum(1 for number in numbers if number % 2 == 0)


def range_band_counts(numbers: Iterable[int]) -> list[int]:
    counts = [0, 0, 0, 0, 0, 0]
    for number in numbers:
        if 1 <= number <= 60:
            counts[(number - 1) // 10] += 1
    return counts


def count_occupied_range_bands(numbers: Iterable[int]) -> int:
    return sum(1 for count in range_band_counts(numbers) if count)


def max_range_band_count(numbers: Iterable[int]) -> int:
    return max(range_band_counts(numbers), default=0)


def _format_int(value: int) -> str:
    return f"{value:,}".replace(",", ".")


def format_int(value: int) -> str:
    """Formata inteiro com separador de milhar brasileiro (ponto)."""
    return _format_int(value)


def _format_percent(value: float) -> str:
    return f"{value:.8f}".rstrip("0").rstrip(".").replace(".", ",")


def format_percent(value: float) -> str:
    """Formata percentual com precisão de 8 casas, usando vírgula decimal."""
    return _format_percent(value)


def draw_parameters(numbers: Iterable[int]) -> dict[str, int]:
    values = list(numbers)
    return {
        "total_sum": sum(values),
        "even_count": count_even_numbers(values),
        "consecutive_count": count_consecutive_numbers(values),
    }


def build_stats() -> dict:
    draw_records = Draw.query.order_by(Draw.contest).all()
    draws = [draw.numbers for draw in draw_records]
    total = len(draws)
    flat = [n for draw in draws for n in draw]
    freq = Counter(flat)
    for n in range(1, 61):
        freq.setdefault(n, 0)

    pair_counter = Counter()
    trio_counter = Counter()
    sums, even_counts = [], []
    consecutive_counts = []
    last_seen = {n: None for n in range(1, 61)}
    for idx, draw_record in enumerate(draw_records, start=1):
        numbers = draw_record.numbers
        pair_counter.update(combinations(numbers, 2))
        trio_counter.update(combinations(numbers, 3))
        sums.append(draw_record.total_sum)
        even_counts.append(draw_record.even_count)
        consecutive_counts.append(draw_record.consecutive_count)
        for n in numbers:
            last_seen[n] = idx

    delays = {n: total - last_seen[n] if last_seen[n] else total for n in range(1, 61)}
    ranges = {"01-10": 0, "11-20": 0, "21-30": 0, "31-40": 0, "41-50": 0, "51-60": 0}
    for n in flat:
        start = ((n - 1) // 10) * 10 + 1
        ranges[f"{start:02d}-{start+9:02d}"] += 1

    sum_distribution = dict(sorted(Counter(sums).items()))
    sum_histogram = _build_sum_histogram(sums)
    even_distribution = dict(sorted(Counter(even_counts).items()))
    consecutive_distribution = dict(sorted(Counter(consecutive_counts).items()))

    prize_cards = {
        "mega_sena": {
            "label": "Mega Sena",
            "games": sum(1 for draw in draw_records if draw.winners_6 > 0),
            "winners": sum(draw.winners_6 for draw in draw_records),
        },
        "quina": {
            "label": "Quina",
            "games": sum(1 for draw in draw_records if draw.winners_5 > 0),
            "winners": sum(draw.winners_5 for draw in draw_records),
        },
        "quadra": {
            "label": "Quadra",
            "games": sum(1 for draw in draw_records if draw.winners_4 > 0),
            "winners": sum(draw.winners_4 for draw in draw_records),
        },
    }
    mega_sena_games_with_winners = prize_cards["mega_sena"]["games"]
    mega_sena_games_without_winners = total - mega_sena_games_with_winners
    mega_sena_games_with_winners_pct = round((mega_sena_games_with_winners / total) * 100, 1) if total else 0
    mega_sena_games_without_winners_pct = round((mega_sena_games_without_winners / total) * 100, 1) if total else 0

    return {
        "total_draws": total,
        "mega_sena_games_with_winners": mega_sena_games_with_winners,
        "mega_sena_games_without_winners": mega_sena_games_without_winners,
        "mega_sena_games_with_winners_pct": mega_sena_games_with_winners_pct,
        "mega_sena_games_without_winners_pct": mega_sena_games_without_winners_pct,
        "frequency": dict(sorted(freq.items())),
        "most_frequent": freq.most_common(10),
        "least_frequent": sorted(freq.items(), key=lambda kv: (kv[1], kv[0]))[:10],
        "top_pairs": pair_counter.most_common(10),
        "top_trios": trio_counter.most_common(10),
        "delays": sorted(delays.items(), key=lambda kv: (-kv[1], kv[0]))[:15],
        "ranges": ranges,
        "avg_sum": round(float(np.mean(sums)), 2) if sums else 0,
        "avg_even": round(float(np.mean(even_counts)), 2) if even_counts else 0,
        "sum_distribution": sum_distribution,
        "sum_histogram": sum_histogram,
        "even_distribution": even_distribution,
        "consecutive_distribution": consecutive_distribution,
        "last_contest": draw_records[-1] if draw_records else None,
        "prize_cards": prize_cards,
    }



def _build_sum_histogram(sums: list[int], bin_size: int = 10) -> dict:
    if not sums:
        return {"bins": [], "max_frequency": 0, "y_ticks": [0]}

    first_bin = (min(sums) // bin_size) * bin_size
    last_bin = math.ceil((max(sums) + 1) / bin_size) * bin_size
    counter = Counter(((total - first_bin) // bin_size) * bin_size + first_bin for total in sums)
    max_frequency = max(counter.values()) if counter else 0
    tick_step = max(1, math.ceil(max_frequency / 4 / 10) * 10)
    y_ticks = list(range(0, tick_step * 5, tick_step))
    scale_max = y_ticks[-1]

    bins = []
    for start in range(first_bin, last_bin, bin_size):
        end = start + bin_size - 1
        bins.append(
            {
                "start": start,
                "end": end,
                "count": counter.get(start, 0),
                "x_label": start if start % 50 == 0 else "",
            }
        )
    return {"bins": bins, "max_frequency": scale_max, "y_ticks": y_ticks}


def count_consecutive_numbers(numbers: Iterable[int]) -> int:
    ordered = sorted(set(numbers))
    longest = 0
    current = 1
    for previous, current_number in zip(ordered, ordered[1:]):
        if current_number == previous + 1:
            current += 1
            longest = max(longest, current)
        else:
            current = 1
    return longest


def refresh_draw_parameters() -> int:
    total_draws = Draw.query.count()
    if total_draws == 0:
        return 0

    updated = 0
    for draw in Draw.query.all():
        derived = draw_parameters(draw.numbers)
        changed = False
        for key, value in derived.items():
            if getattr(draw, key) != value:
                setattr(draw, key, value)
                changed = True
        if changed:
            updated += 1
    if updated:
        db.session.commit()
        _log.info("refresh_draw_parameters: %d/%d concursos recalculados.", updated, total_draws)
    return updated


@lru_cache(maxsize=1)
def _combination_distribution() -> dict[tuple[int, int, int, int, int], int]:
    states: dict[tuple[int, int, int, int, int, int, int, int], int] = {(0, 0, 0, 0, 0, 0, 0, 0): 1}
    for number in range(1, 61):
        if number in {11, 21, 31, 41, 51}:
            reset_states: dict[tuple[int, int, int, int, int, int, int, int], int] = {}
            for (quantity, total_sum, even_count, current_run, longest_run, occupied_bands, _current_band_count, max_band_count), count in states.items():
                key = (quantity, total_sum, even_count, current_run, longest_run, occupied_bands, 0, max_band_count)
                reset_states[key] = reset_states.get(key, 0) + count
            states = reset_states

        next_states: dict[tuple[int, int, int, int, int, int, int, int], int] = {}
        for (quantity, total_sum, even_count, current_run, longest_run, occupied_bands, current_band_count, max_band_count), count in states.items():
            skip_key = (quantity, total_sum, even_count, 0, longest_run, occupied_bands, current_band_count, max_band_count)
            next_states[skip_key] = next_states.get(skip_key, 0) + count

            if quantity < 6:
                new_current_run = current_run + 1 if current_run else 1
                new_longest_run = max(longest_run, new_current_run if new_current_run >= 2 else 0)
                new_current_band_count = current_band_count + 1
                new_occupied_bands = occupied_bands + (1 if current_band_count == 0 else 0)
                new_max_band_count = max(max_band_count, new_current_band_count)
                take_key = (
                    quantity + 1,
                    total_sum + number,
                    even_count + (1 if number % 2 == 0 else 0),
                    new_current_run,
                    new_longest_run,
                    new_occupied_bands,
                    new_current_band_count,
                    new_max_band_count,
                )
                next_states[take_key] = next_states.get(take_key, 0) + count
        states = next_states

    distribution: dict[tuple[int, int, int, int, int], int] = {}
    for (quantity, total_sum, even_count, _current_run, longest_run, occupied_bands, _current_band_count, max_band_count), count in states.items():
        if quantity == 6:
            key = (longest_run, even_count, total_sum, occupied_bands, max_band_count)
            distribution[key] = distribution.get(key, 0) + count
    return distribution


def count_possible_draw_combinations(
    consecutive_count: int | None = None,
    even_min: int | None = None,
    even_max: int | None = None,
    sum_min: int | None = None,
    sum_max: int | None = None,
    range_min_occupied: int | None = None,
    range_max_per_band: int | None = None,
) -> int:
    total = 0
    for (longest_run, even_numbers, total_sum, occupied_bands, max_band_count), count in _combination_distribution().items():
        if consecutive_count is not None and longest_run > consecutive_count:
            continue
        if even_min is not None and even_numbers < even_min:
            continue
        if even_max is not None and even_numbers > even_max:
            continue
        if sum_min is not None and total_sum < sum_min:
            continue
        if sum_max is not None and total_sum > sum_max:
            continue
        if range_min_occupied is not None and occupied_bands < range_min_occupied:
            continue
        if range_max_per_band is not None and max_band_count > range_max_per_band:
            continue
        total += count
    return total


def build_combination_report(quantity: int = 6, filters: dict | None = None) -> dict:
    filters = filters or {}
    quantity = max(6, min(int(quantity), 15))
    total = math.comb(60, 6)
    remaining = total
    steps = []

    filter_steps = [
        ("even", "Quantidade de números pares", (filters.get("even_min"), filters.get("even_max"))),
        ("sum", "Soma dos números", (filters.get("sum_min"), filters.get("sum_max"))),
        ("range", "Distribuição por faixas", (filters.get("range_min_occupied"), filters.get("range_max_per_band"))),
        ("consecutive_count", "Maior sequência de números consecutivos", filters.get("consecutive_count")),
    ]

    active_filters: dict[str, int] = {}
    for key, label, value in filter_steps:
        step_number = len(steps) + 1
        if key == "sum":
            value_min, value_max = value
            if value_min is None and value_max is None:
                continue
            lower_bound = value_min if value_min is not None else 21
            upper_bound = value_max if value_max is not None else 345
            if value_min is not None:
                active_filters["sum_min"] = value_min
            if value_max is not None:
                active_filters["sum_max"] = value_max
            display_value = f"{lower_bound} a {upper_bound}"
            formula = f"S{step_number} = {{ jogo em S{step_number - 1} | {lower_bound} <= soma(jogo) <= {upper_bound} }}"
            explanation = "A soma de cada jogo é comparada com o intervalo informado. Jogos abaixo da soma inicial ou acima da soma final saem do universo restante."
        elif key == "even":
            value_min, value_max = value
            if value_min is None and value_max is None:
                continue
            lower_bound = value_min if value_min is not None else 0
            upper_bound = value_max if value_max is not None else 6
            if value_min is not None:
                active_filters["even_min"] = value_min
            if value_max is not None:
                active_filters["even_max"] = value_max
            display_value = f"{lower_bound} a {upper_bound}"
            formula = f"S{step_number} = {{ jogo em S{step_number - 1} | {lower_bound} <= quantidade_de_pares(jogo) <= {upper_bound} }}"
            explanation = "Conta os números divisíveis por 2 dentro de cada jogo. O filtro mantém todos os jogos cuja quantidade de pares fique dentro da faixa informada."
        elif key == "range":
            min_occupied, max_per_band = value
            if min_occupied is None and max_per_band is None:
                continue
            if min_occupied is not None:
                active_filters["range_min_occupied"] = min_occupied
            if max_per_band is not None:
                active_filters["range_max_per_band"] = max_per_band
            parts = []
            formula_parts = []
            if min_occupied is not None:
                parts.append(f"mín. {min_occupied} faixas")
                formula_parts.append(f"faixas_ocupadas(jogo) >= {min_occupied}")
            if max_per_band is not None:
                parts.append(f"máx. {max_per_band} por faixa")
                formula_parts.append(f"max_por_faixa(jogo) <= {max_per_band}")
            display_value = ", ".join(parts)
            formula = f"S{step_number} = {{ jogo em S{step_number - 1} | {' e '.join(formula_parts)} }}"
            explanation = "Divide as dezenas nos blocos 01-10, 11-20, 21-30, 31-40, 41-50 e 51-60. O filtro controla quantas faixas precisam aparecer e quantas dezenas podem ficar concentradas na mesma faixa."
        else:
            if value is None:
                continue
            active_filters[key] = value
            display_value = f"até {value}"
            formula = f"S{step_number} = {{ jogo em S{step_number - 1} | maior_sequencia_consecutiva(jogo) <= {value} }}"
            explanation = "Ordena os 6 números do jogo e mede o maior bloco em que cada número vem logo depois do anterior. O filtro mantém jogos com sequência máxima menor ou igual ao limite informado."

        new_remaining = count_possible_draw_combinations(**active_filters)
        eliminated = remaining - new_remaining
        steps.append(
            {
                "label": label,
                "value": display_value,
                "formula": formula,
                "explanation": explanation,
                "previous_remaining": remaining,
                "previous_remaining_formatted": _format_int(remaining),
                "eliminated": eliminated,
                "remaining": new_remaining,
                "eliminated_formatted": _format_int(eliminated),
                "remaining_formatted": _format_int(new_remaining),
            }
        )
        remaining = new_remaining

    covered_combinations = math.comb(quantity, 6)
    chance = min(covered_combinations / remaining, 1.0) if remaining else 0
    return {
        "total": total,
        "total_formatted": _format_int(total),
        "remaining": remaining,
        "remaining_formatted": _format_int(remaining),
        "eliminated": total - remaining,
        "eliminated_formatted": _format_int(total - remaining),
        "covered_combinations": covered_combinations,
        "covered_combinations_formatted": _format_int(covered_combinations),
        "chance_percent": chance * 100,
        "chance_percent_formatted": _format_percent(chance * 100),
        "chance_one_in": math.ceil(1 / chance) if chance else None,
        "chance_one_in_formatted": _format_int(math.ceil(1 / chance)) if chance else "0",
        "steps": steps,
    }


def count_draws_matching_filters(
    consecutive_count: int | None = None,
    even_min: int | None = None,
    even_max: int | None = None,
    sum_min: int | None = None,
    sum_max: int | None = None,
    range_min_occupied: int | None = None,
    range_max_per_band: int | None = None,
) -> int:
    query = Draw.query
    if consecutive_count is not None:
        query = query.filter(Draw.consecutive_count <= consecutive_count)
    if even_min is not None:
        query = query.filter(Draw.even_count >= even_min)
    if even_max is not None:
        query = query.filter(Draw.even_count <= even_max)
    if sum_min is not None:
        query = query.filter(Draw.total_sum >= sum_min)
    if sum_max is not None:
        query = query.filter(Draw.total_sum <= sum_max)
    if range_min_occupied is None and range_max_per_band is None:
        return query.count()
    count = 0
    for draw in query.all():
        numbers = draw.numbers
        if range_min_occupied is not None and count_occupied_range_bands(numbers) < range_min_occupied:
            continue
        if range_max_per_band is not None and max_range_band_count(numbers) > range_max_per_band:
            continue
        count += 1
    return count


def calculate_individual_filter_targets(target_percentage: float) -> dict:
    draws = Draw.query.with_entities(
        Draw.consecutive_count,
        Draw.even_count,
        Draw.total_sum,
        Draw.n1,
        Draw.n2,
        Draw.n3,
        Draw.n4,
        Draw.n5,
        Draw.n6,
    ).all()
    total = len(draws)
    try:
        target_percentage = float(target_percentage)
    except (TypeError, ValueError):
        target_percentage = 80.0
    if not math.isfinite(target_percentage):
        target_percentage = 80.0
    target_percentage = max(0.0, min(target_percentage, 100.0))
    target_count = math.ceil((target_percentage / 100) * total) if total else 0

    def metric(value: int | None, count: int = 0) -> dict:
        percentage = (count / total) * 100 if total else 0
        return {
            "value": value,
            "count": count,
            "percentage": round(percentage, 2),
            "percentage_text": f"{percentage:.2f}".replace(".", ",") + "%",
        }

    if not total:
        empty_parameters = {
            "consecutive_count": metric(None),
            "even_min": metric(None),
            "even_max": metric(None),
            "sum_min": metric(None),
            "sum_max": metric(None),
            "range_min_occupied": metric(None),
            "range_max_per_band": metric(None),
        }
        return {"target_percentage": target_percentage, "total": 0, "target_count": 0, "parameters": empty_parameters}

    consecutive_values = [row.consecutive_count for row in draws]
    even_values = [row.even_count for row in draws]
    sum_values = [row.total_sum for row in draws]
    draw_numbers = [[row.n1, row.n2, row.n3, row.n4, row.n5, row.n6] for row in draws]
    occupied_range_values = [count_occupied_range_bands(numbers) for numbers in draw_numbers]
    max_range_values = [max_range_band_count(numbers) for numbers in draw_numbers]

    def first_at_or_above(candidates: Iterable[int], counter) -> dict:
        for value in candidates:
            count = counter(value)
            if count >= target_count:
                return metric(value, count)
        value = list(candidates)[-1]
        return metric(value, counter(value))

    def last_at_or_above(candidates: Iterable[int], counter) -> dict:
        selected_value = None
        selected_count = 0
        for value in candidates:
            count = counter(value)
            if count >= target_count:
                selected_value = value
                selected_count = count
        return metric(selected_value, selected_count)

    unique_sums = sorted(set(sum_values))
    parameters = {
        "consecutive_count": first_at_or_above(range(0, 7), lambda value: sum(1 for item in consecutive_values if item <= value)),
        "even_min": last_at_or_above(range(0, 7), lambda value: sum(1 for item in even_values if item >= value)),
        "even_max": first_at_or_above(range(0, 7), lambda value: sum(1 for item in even_values if item <= value)),
        "sum_min": last_at_or_above(unique_sums, lambda value: sum(1 for item in sum_values if item >= value)),
        "sum_max": first_at_or_above(unique_sums, lambda value: sum(1 for item in sum_values if item <= value)),
        "range_min_occupied": last_at_or_above(range(1, 7), lambda value: sum(1 for item in occupied_range_values if item >= value)),
        "range_max_per_band": first_at_or_above(range(1, 7), lambda value: sum(1 for item in max_range_values if item <= value)),
    }
    return {
        "target_percentage": target_percentage,
        "total": total,
        "target_count": target_count,
        "parameters": parameters,
    }


def _passes_generation_filters(numbers: list[int], filters: dict | None) -> bool:
    if not filters:
        return True
    consecutive_count = filters.get("consecutive_count")
    even_min = filters.get("even_min")
    even_max = filters.get("even_max")
    sum_min = filters.get("sum_min")
    sum_max = filters.get("sum_max")
    range_min_occupied = filters.get("range_min_occupied")
    range_max_per_band = filters.get("range_max_per_band")
    total_sum = sum(numbers)
    even_count = count_even_numbers(numbers)
    if even_min is not None and even_count < even_min:
        return False
    if even_max is not None and even_count > even_max:
        return False
    if sum_min is not None and total_sum < sum_min:
        return False
    if sum_max is not None and total_sum > sum_max:
        return False
    if consecutive_count is not None and count_consecutive_numbers(numbers) > consecutive_count:
        return False
    if range_min_occupied is not None and count_occupied_range_bands(numbers) < range_min_occupied:
        return False
    if range_max_per_band is not None and max_range_band_count(numbers) > range_max_per_band:
        return False
    return True

def _diversity_score(numbers: list[int], existing_candidates: list[list[int]]) -> float:
    if not existing_candidates:
        return 1.0
    max_overlap = max(len(set(numbers) & set(candidate)) for candidate in existing_candidates)
    return max(0.0, 1.0 - (max_overlap / max(len(numbers), 1)))


def _passes_diversity_control(numbers: list[int], created_numbers: list[list[int]]) -> bool:
    if not created_numbers:
        return True
    # Evita apostas praticamente iguais dentro da mesma geração.
    # Para apostas de 6 dezenas, no máximo 4 números podem se repetir entre duas apostas.
    max_allowed_overlap = max(0, len(numbers) - 2)
    current = set(numbers)
    return all(len(current & set(candidate)) <= max_allowed_overlap for candidate in created_numbers)


def _secure_random_candidate(quantity: int) -> list[int]:
    rng = secrets.SystemRandom()
    return sorted(rng.sample(range(1, 61), quantity))


def generate_bets(
    quantity: int,
    amount: int,
    persist: bool = True,
    filters: dict | None = None,
) -> list[GeneratedBet]:
    draws = all_draw_numbers()
    existing_draws = {tuple(d) for d in draws}
    created: list[GeneratedBet] = []
    created_numbers: list[list[int]] = []
    attempts = 0
    # Limite razoável: evita loop infinito com filtros muito restritivos.
    # 2000 tentativas por aposta é suficiente para qualquer configuração prática.
    max_attempts = amount * 2000
    while len(created) < amount and attempts < max_attempts:
        attempts += 1
        nums = _secure_random_candidate(quantity)
        if quantity == 6 and tuple(nums) in existing_draws:
            continue
        if not _passes_generation_filters(nums, filters):
            continue
        if not _passes_diversity_control(nums, created_numbers):
            continue
        score = _diversity_score(nums, created_numbers)
        bet = GeneratedBet(
            quantity=quantity,
            numbers_csv=",".join(map(str, nums)),
            score=round(score, 4),
        )
        if persist:
            db.session.add(bet)
        created.append(bet)
        created_numbers.append(nums)
    if len(created) < amount:
        _log.warning(
            "generate_bets: geradas %d/%d apostas após %d tentativas (filtros: %s).",
            len(created), amount, attempts, filters,
        )
    if persist:
        db.session.commit()
    return created


def generate_closure_bets(numbers: Iterable[int]) -> list[GeneratedBet]:
    base_numbers = sorted(set(numbers))
    if len(base_numbers) < 6:
        raise RuntimeError("Informe pelo menos 6 dezenas distintas para gerar um fechamento matemático.")
    if len(base_numbers) > 15:
        raise RuntimeError("Use no máximo 15 dezenas no fechamento matemático.")
    if any(number < 1 or number > 60 for number in base_numbers):
        raise RuntimeError("As dezenas do fechamento devem estar entre 1 e 60.")

    return [
        GeneratedBet(
            quantity=6,
            numbers_csv=",".join(map(str, combination)),
            score=0,
        )
        for combination in combinations(base_numbers, 6)
    ]

def build_recent_frequency(count: int | None) -> dict:
    """
    Frequência de aparição de cada dezena nos últimos `count` concursos.
    Se `count` for None, considera todos os concursos históricos.
    """
    query = Draw.query.order_by(Draw.contest.desc())
    if count is not None:
        query = query.limit(count)
    draws = query.all()

    if not draws:
        return {
            "count": count,
            "actual_count": 0,
            "frequency": {str(n): 0 for n in range(1, 61)},
            "max_frequency": 0,
            "most_frequent": [],
            "least_frequent": [],
        }

    flat = [n for draw in draws for n in draw.numbers]
    freq: Counter[int] = Counter(flat)
    for n in range(1, 61):
        freq.setdefault(n, 0)

    return {
        "count": count,
        "actual_count": len(draws),
        "frequency": {str(n): freq[n] for n in range(1, 61)},
        "max_frequency": max(freq.values()),
        "most_frequent": freq.most_common(10),
        "least_frequent": sorted(freq.items(), key=lambda kv: (kv[1], kv[0]))[:10],
    }


def list_recent_generations(limit: int = 12) -> list[dict]:
    rows = (
        db.session.query(
            GeneratedBet.generation_id,
            func.count(GeneratedBet.id),
            func.min(GeneratedBet.quantity),
            func.max(GeneratedBet.created_at),
        )
        .filter(GeneratedBet.generation_id.isnot(None))
        .group_by(GeneratedBet.generation_id)
        .order_by(func.max(GeneratedBet.created_at).desc(), GeneratedBet.generation_id.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "generation_id": generation_id,
            "bet_count": bet_count,
            "quantity": quantity,
            "created_at": created_at,
        }
        for generation_id, bet_count, quantity, created_at in rows
    ]


def list_recent_generations_with_bets(limit: int = 12) -> list[dict]:
    generations = list_recent_generations(limit=limit)
    generation_ids = [generation["generation_id"] for generation in generations]
    if not generation_ids:
        return generations

    bets = (
        GeneratedBet.query.filter(GeneratedBet.generation_id.in_(generation_ids))
        .order_by(GeneratedBet.generation_id.desc(), GeneratedBet.id)
        .all()
    )
    bets_by_generation: dict[int, list[GeneratedBet]] = {}
    for bet in bets:
        bets_by_generation.setdefault(bet.generation_id, []).append(bet)

    for generation in generations:
        generation["bets"] = bets_by_generation.get(generation["generation_id"], [])
    return generations


def save_generated_bets(quantity: int, bets: Iterable[str]) -> tuple[int, int | None]:
    valid_bets = []
    for numbers_csv in bets:
        nums = [_to_int(n) for n in numbers_csv.split(",")]
        if len(nums) != quantity or any(n is None or n < 1 or n > 60 for n in nums) or len(set(nums)) != quantity:
            continue
        nums = sorted(nums)  # type: ignore[arg-type]
        valid_bets.append(",".join(map(str, nums)))

    if not valid_bets:
        return 0, None

    # Usa SELECT dentro da mesma transação para minimizar corrida no generation_id.
    # Em SQLite com WAL e uso single-user isso é suficiente; não requer SERIALIZABLE.
    last_generation_id = db.session.query(func.max(GeneratedBet.generation_id)).scalar() or 0
    generation_id = last_generation_id + 1
    for numbers_csv in valid_bets:
        db.session.add(
            GeneratedBet(generation_id=generation_id, quantity=quantity, numbers_csv=numbers_csv, score=0)
        )
    db.session.commit()
    _log.info("Apostas salvas: %d na geração #%d.", len(valid_bets), generation_id)
    return len(valid_bets), generation_id
