from __future__ import annotations

import logging
import unicodedata
from collections.abc import Iterable
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import BinaryIO
from zipfile import BadZipFile, ZipFile

from ..core.numbers import MAX_INT32, MAX_INT64, draw_parameters, parse_int
from ..extensions import db
from ..models import Draw

_log = logging.getLogger(__name__)
MAX_IMPORT_ROWS = 10_000
MAX_XLSX_ARCHIVE_FILES = 1_000
MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 200


def _norm(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return " ".join(text.lower().replace("_", " ").replace("-", " ").split())


def _parse_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        raise ValueError("data inválida")
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError("data inválida")


def _parse_winner_count(value: object) -> int:
    if value is None or value == "":
        return 0
    parsed = parse_int(value, max_abs=MAX_INT32)
    if parsed is None or parsed < 0:
        raise ValueError("quantidade de ganhadores inválida")
    return parsed


def _count_rows_with_limit(rows: Iterable[tuple[object, ...]]) -> int:
    count = 0
    for count, _row in enumerate(rows, start=1):
        if count > MAX_IMPORT_ROWS:
            raise RuntimeError(
                f"A planilha ultrapassa o limite de {MAX_IMPORT_ROWS} linhas de dados."
            )
    return count


def _money_to_cents(value: object) -> int:
    """Converte um valor monetário explícito para centavos.

    Uma célula vazia representa zero. Já um valor não vazio que não possa ser
    interpretado como dinheiro é um erro da planilha: não o trate como zero,
    pois isso poderia apagar uma premiação já armazenada ao reimportar um
    concurso.
    """
    if value is None or value == "":
        return 0
    if isinstance(value, bool):
        raise ValueError("valor monetário inválido")
    try:
        if isinstance(value, (int, float, Decimal)):
            amount = Decimal(str(value))
        else:
            text = (
                str(value)
                .strip()
                .replace("R$", "")
                .replace("\u00a0", "")
                .replace(" ", "")
            )
            if not text or len(text) > 64:
                raise ValueError("valor monetário inválido")
            if "," in text and "." in text:
                text = (
                    text.replace(".", "").replace(",", ".")
                    if text.rfind(",") > text.rfind(".")
                    else text.replace(",", "")
                )
            elif "," in text:
                text = text.replace(",", ".")
            amount = Decimal(text)
        if not amount.is_finite() or amount < 0:
            raise ValueError("valor monetário inválido")
        cents = int((amount * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        if cents > MAX_INT64:
            raise ValueError("valor monetário fora do limite aceito")
        return cents
    except (InvalidOperation, OverflowError, TypeError, ValueError):
        raise ValueError("valor monetário inválido") from None


def _validate_xlsx_archive(source: str | Path | BinaryIO) -> None:
    """Rejeita arquivos XLSX corrompidos, criptografados ou desproporcionalmente expandidos."""
    original_position = None
    if hasattr(source, "tell") and hasattr(source, "seek"):
        try:
            original_position = source.tell()
            source.seek(0)
        except (OSError, ValueError):
            original_position = None
    try:
        with ZipFile(source) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_ARCHIVE_FILES:
                raise RuntimeError("A planilha contém arquivos internos demais.")
            if any(member.flag_bits & 0x1 for member in members):
                raise RuntimeError("Planilhas XLSX criptografadas não são suportadas.")
            total_size = sum(member.file_size for member in members)
            compressed_size = sum(member.compress_size for member in members)
            if total_size > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise RuntimeError(
                    "A planilha é grande demais depois de descompactada."
                )
            ratio = total_size / max(compressed_size, 1)
            if ratio > MAX_XLSX_COMPRESSION_RATIO:
                raise RuntimeError("A planilha possui uma taxa de compressão insegura.")
    except BadZipFile as exc:
        raise RuntimeError(
            "Não foi possível ler o arquivo: planilha XLSX inválida."
        ) from exc
    finally:
        if original_position is not None:
            source.seek(original_position)


def import_results_from_xlsx(source: str | Path | BinaryIO) -> dict[str, int]:
    _validate_xlsx_archive(source)
    try:
        # openpyxl tem custo de importacao relevante; carregue-o apenas quando
        # o usuario realmente importar uma planilha.
        from openpyxl import load_workbook

        workbook = load_workbook(
            source, read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:
        _log.error("Falha ao abrir planilha: %s", exc)
        raise RuntimeError(f"Não foi possível ler o arquivo: {exc}") from exc

    try:
        sheet = workbook[workbook.sheetnames[0]]
        if hasattr(sheet, "reset_dimensions"):
            sheet.reset_dimensions()
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return {"imported": 0, "updated": 0, "ignored": 0}

        normalized = [_norm(c) for c in header]

        def find_one(candidates: Iterable[str]) -> int | None:
            cand = {_norm(c) for c in candidates}
            for idx, name in enumerate(normalized):
                if name in cand:
                    return idx
            for idx, name in enumerate(normalized):
                if any(c in name for c in cand):
                    return idx
            return None

        contest_idx = find_one(["concurso", "contest", "numero concurso", "n concurso"])
        date_idx = find_one(["data sorteio", "data", "draw date"])
        winners_6_idx = find_one(["ganhadores 6 acertos", "ganhadores sena", "sena"])
        winners_5_idx = find_one(["ganhadores 5 acertos", "ganhadores quina", "quina"])
        winners_4_idx = find_one(
            ["ganhadores 4 acertos", "ganhadores quadra", "quadra"]
        )
        prize_idx = find_one(["rateio 6 acertos", "premio", "prêmio"])
        accumulated_idx = find_one(["acumulado 6 acertos", "acumulado"])
        quina_rateio_idx = find_one(["rateio 5 acertos", "rateio quina"])
        quadra_rateio_idx = find_one(["rateio 4 acertos", "rateio quadra"])

        number_indexes: list[int] = []
        for token in [
            "bola 1",
            "bola1",
            "dezena 1",
            "n1",
            "bola 2",
            "bola2",
            "dezena 2",
            "n2",
            "bola 3",
            "bola3",
            "dezena 3",
            "n3",
            "bola 4",
            "bola4",
            "dezena 4",
            "n4",
            "bola 5",
            "bola5",
            "dezena 5",
            "n5",
            "bola 6",
            "bola6",
            "dezena 6",
            "n6",
        ]:
            idx = find_one([token])
            if idx is not None and idx not in number_indexes:
                number_indexes.append(idx)
            if len(number_indexes) == 6:
                break

        if contest_idx is None or len(number_indexes) < 6:
            return {
                "imported": 0,
                "updated": 0,
                "ignored": _count_rows_with_limit(rows),
            }

        imported = updated = ignored = 0
        existing_draws = {draw.contest: draw for draw in Draw.query.all()}
        seen_contests: set[int] = set()
        try:
            for row_number, row in enumerate(rows, start=1):
                if row_number > MAX_IMPORT_ROWS:
                    raise RuntimeError(
                        f"A planilha ultrapassa o limite de {MAX_IMPORT_ROWS} linhas de dados."
                    )
                contest = parse_int(
                    row[contest_idx] if contest_idx < len(row) else None,
                    max_abs=MAX_INT32,
                )
                numbers = [
                    parse_int(row[i] if i < len(row) else None)
                    for i in number_indexes[:6]
                ]
                if (
                    contest is None
                    or contest <= 0
                    or any(n is None or n < 1 or n > 60 for n in numbers)
                    or len(set(numbers)) != 6
                ):
                    ignored += 1
                    continue
                if contest in seen_contests:
                    ignored += 1
                    continue
                numbers = sorted(numbers)  # type: ignore[arg-type]
                derived = draw_parameters(numbers)
                payload = {
                    "n1": numbers[0],
                    "n2": numbers[1],
                    "n3": numbers[2],
                    "n4": numbers[3],
                    "n5": numbers[4],
                    "n6": numbers[5],
                    "total_sum": derived["total_sum"],
                    "even_count": derived["even_count"],
                    "consecutive_count": derived["consecutive_count"],
                }
                optional_payload: dict[str, object] = {}

                # `row` é ligado como padrão em vez de capturado do escopo do
                # laço. Hoje a função só é chamada dentro da própria iteração,
                # então o resultado é o mesmo; a ligação explícita evita que
                # guardá-la para uso posterior passe a ler a linha errada.
                def cell_at(index: int, *, row: list = row) -> object:
                    return row[index] if index < len(row) else None

                if date_idx is not None:
                    try:
                        optional_payload["draw_date"] = _parse_date(cell_at(date_idx))
                    except ValueError as exc:
                        raise RuntimeError(
                            f"Data inválida no concurso {contest}."
                        ) from exc
                for field, index, label in (
                    ("winners_6", winners_6_idx, "ganhadores de 6 acertos"),
                    ("winners_5", winners_5_idx, "ganhadores de 5 acertos"),
                    ("winners_4", winners_4_idx, "ganhadores de 4 acertos"),
                ):
                    if index is not None:
                        try:
                            optional_payload[field] = _parse_winner_count(cell_at(index))
                        except ValueError as exc:
                            raise RuntimeError(
                                f"Quantidade inválida no concurso {contest} ({label})."
                            ) from exc
                for field, index, label in (
                    ("prize_cents", prize_idx, "rateio de 6 acertos"),
                    ("accumulated_cents", accumulated_idx, "acumulado"),
                    ("quina_rateio_cents", quina_rateio_idx, "rateio de 5 acertos"),
                    ("quadra_rateio_cents", quadra_rateio_idx, "rateio de 4 acertos"),
                ):
                    if index is not None:
                        try:
                            optional_payload[field] = _money_to_cents(cell_at(index))
                        except ValueError as exc:
                            raise RuntimeError(
                                f"Valor monetário inválido no concurso {contest} ({label})."
                            ) from exc
                draw = existing_draws.get(contest)
                if draw:
                    payload.update(optional_payload)
                    if all(
                        getattr(draw, key) == value for key, value in payload.items()
                    ):
                        ignored += 1
                        seen_contests.add(contest)
                        continue
                    for key, value in payload.items():
                        setattr(draw, key, value)
                    updated += 1
                else:
                    draw = Draw(contest=contest, **payload, **optional_payload)
                    db.session.add(draw)
                    existing_draws[contest] = draw
                    imported += 1
                seen_contests.add(contest)
            db.session.commit()
        except Exception:
            db.session.rollback()
            raise
        _log.info(
            "Importação concluída: %d novos, %d atualizados, %d ignorados.",
            imported,
            updated,
            ignored,
        )
        return {"imported": imported, "updated": updated, "ignored": ignored}
    finally:
        workbook.close()
